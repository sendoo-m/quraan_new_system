from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views import View
from django.views.generic.edit import UpdateView
from django.urls import reverse_lazy

from accounts.permissions import (
    GeneralSupervisorRequiredMixin,
    StaffRequiredMixin,
    user_can_access_hall,
)
from accounts.models import GeneralSupervisorHallAssignment
from .models import Hall, Subject, HallSchedule, ScheduleTemplate, ScheduleTemplateEntry, ALL_DAYS
from .forms import HallForm, HallScheduleForm, SubjectForm, ScheduleTemplateForm, ScheduleTemplateEntryForm

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════
#  دوال مساعدة
# ══════════════════════════════════════════

def get_general_supervisor_names(hall):
    """
    إرجاع أسماء كل المشرفين العامين المسندين للقاعة.
    مع fallback للحقل القديم general_supervisor لو موجود.
    """
    names = []

    if hasattr(hall, 'general_supervisor_assignments'):
        for assignment in hall.general_supervisor_assignments.all():
            if assignment.supervisor:
                names.append(
                    assignment.supervisor.get_full_name()
                    or assignment.supervisor.username
                )

    # توافق مع النظام القديم لو لم توجد إسنادات في جدول GeneralSupervisorHallAssignment
    if not names and hasattr(hall, 'general_supervisor') and hall.general_supervisor:
        names.append(
            hall.general_supervisor.get_full_name()
            or hall.general_supervisor.username
        )

    return '، '.join(names) if names else '—'


def save_hall_general_supervisors(form, hall, request_user=None):
    """
    حفظ إسنادات المشرفين العامين المتعددة للقاعة.
    يدعم HallForm لو فيه save_general_supervisors،
    ويدعم الحفظ اليدوي كـ fallback.
    """
    if hasattr(form, 'save_general_supervisors'):
        form.save_general_supervisors(hall)
        return

    selected_supervisors = form.cleaned_data.get('general_supervisors')

    if request_user and request_user.is_general_supervisor and not request_user.is_general_manager:
        selected_supervisors = [request_user]

    GeneralSupervisorHallAssignment.objects.filter(hall=hall).delete()

    if selected_supervisors:
        for supervisor in selected_supervisors:
            GeneralSupervisorHallAssignment.objects.get_or_create(
                hall=hall,
                supervisor=supervisor
            )


# ══════════════════════════════════════════
#  القاعات
# ══════════════════════════════════════════

class HallListView(StaffRequiredMixin, View):
    def get_queryset(self, request):
        user = request.user

        qs = Hall.objects.select_related(
            'teacher',
            'supervisor',
            'age_group',
        ).prefetch_related(
            'general_supervisor_assignments__supervisor'
        )

        if user.is_general_manager:
            return qs

        if user.is_general_supervisor:
            return qs.filter(
                general_supervisor_assignments__supervisor=user
            ).distinct()

        if user.is_hall_supervisor:
            return qs.filter(supervisor=user)

        if user.is_teacher:
            return qs.filter(teacher=user)

        return qs.none()

    def get(self, request):
        halls = self.get_queryset(request)
        return render(request, 'halls/list.html', {
            'halls': halls,
            'total': halls.count(),
            'active': halls.filter(is_active=True).count(),
        })


class HallCreateView(GeneralSupervisorRequiredMixin, View):
    template_name = 'halls/form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'form': HallForm(user=request.user),
            'action': 'إضافة'
        })

    def post(self, request):
        form = HallForm(request.POST, user=request.user)

        if form.is_valid():
            hall = form.save(commit=False)

            # توافق مع الحقل القديم داخل Hall لو ما زال موجودًا
            if hasattr(hall, 'general_supervisor'):
                if request.user.is_general_supervisor and not request.user.is_general_manager:
                    hall.general_supervisor = request.user
                else:
                    selected_supervisors = form.cleaned_data.get('general_supervisors')
                    hall.general_supervisor = selected_supervisors.first() if selected_supervisors else None

            hall.save()
            form.save_m2m()

            save_hall_general_supervisors(form, hall, request.user)

            messages.success(request, f'✅ تم إنشاء قاعة {hall.name} بنجاح')
            return redirect('halls:list')

        return render(request, self.template_name, {
            'form': form,
            'action': 'إضافة'
        })


class HallDetailView(StaffRequiredMixin, View):
    def get(self, request, pk):
        hall = get_object_or_404(
            Hall.objects.select_related(
                'teacher',
                'supervisor',
                'age_group',
            ).prefetch_related(
                'general_supervisor_assignments__supervisor'
            ),
            pk=pk
        )

        if not user_can_access_hall(request.user, hall):
            messages.error(request, 'ليس لديك صلاحية عرض هذه القاعة')
            return redirect('halls:list')

        return render(request, 'halls/detail.html', {
            'hall': hall,
            'students': hall.students.filter(status='active').select_related('parent'),
            'schedules': hall.schedules.select_related('subject').order_by('day', 'start_time'),
            'general_supervisor_names': get_general_supervisor_names(hall),
        })


class HallUpdateView(GeneralSupervisorRequiredMixin, UpdateView):
    model = Hall
    form_class = HallForm
    template_name = 'halls/form.html'
    success_url = reverse_lazy('halls:list')

    def get_queryset(self):
        return Hall.objects.select_related(
            'teacher',
            'supervisor',
            'age_group',
        ).prefetch_related(
            'general_supervisor_assignments__supervisor'
        )

    def dispatch(self, request, *args, **kwargs):
        if not user_can_access_hall(request.user, self.get_object()):
            messages.error(request, 'ليس لديك صلاحية تعديل هذه القاعة')
            return redirect('halls:list')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        hall = form.save(commit=False)

        # توافق مع الحقل القديم داخل Hall لو ما زال موجودًا
        if hasattr(hall, 'general_supervisor'):
            if self.request.user.is_general_supervisor and not self.request.user.is_general_manager:
                hall.general_supervisor = self.request.user
            else:
                selected_supervisors = form.cleaned_data.get('general_supervisors')
                hall.general_supervisor = selected_supervisors.first() if selected_supervisors else None

        hall.save()
        form.save_m2m()

        save_hall_general_supervisors(form, hall, self.request.user)

        messages.success(self.request, f'✅ تم تعديل بيانات القاعة {hall.name}')
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['action'] = 'تعديل'
        return ctx


# ══════════════════════════════════════════
#  جداول القاعات المباشرة legacy
# ══════════════════════════════════════════

class HallScheduleView(GeneralSupervisorRequiredMixin, View):
    template_name = 'halls/schedule.html'

    def get(self, request, pk):
        hall = get_object_or_404(Hall, pk=pk)

        if not user_can_access_hall(request.user, hall):
            messages.error(request, 'ليس لديك صلاحية إدارة جدول هذه القاعة')
            return redirect('halls:list')

        return render(request, self.template_name, {
            'hall': hall,
            'schedules': hall.schedules.select_related('subject'),
            'form': HallScheduleForm(hall=hall),
        })

    def post(self, request, pk):
        hall = get_object_or_404(Hall, pk=pk)

        if not user_can_access_hall(request.user, hall):
            messages.error(request, 'ليس لديك صلاحية إدارة جدول هذه القاعة')
            return redirect('halls:list')

        form = HallScheduleForm(request.POST, hall=hall)

        if form.is_valid():
            schedule = form.save(commit=False)
            schedule.hall = hall
            schedule.save()

            messages.success(request, '✅ تم إضافة الحصة للجدول')
            return redirect('halls:schedule', pk=pk)

        return render(request, self.template_name, {
            'hall': hall,
            'schedules': hall.schedules.select_related('subject'),
            'form': form,
        })


# ══════════════════════════════════════════
#  عرض كل الجداول
# ══════════════════════════════════════════

class AllSchedulesView(StaffRequiredMixin, View):
    def get_allowed_halls(self, request):
        user = request.user

        qs = Hall.objects.filter(is_active=True).select_related(
            'teacher',
            'supervisor',
            'age_group',
            'schedule_template',
        ).prefetch_related(
            'general_supervisor_assignments__supervisor'
        )

        if user.is_general_manager:
            return qs

        if user.is_general_supervisor:
            return qs.filter(
                general_supervisor_assignments__supervisor=user
            ).distinct()

        if user.is_hall_supervisor:
            return qs.filter(supervisor=user)

        if user.is_teacher:
            return qs.filter(teacher=user)

        return qs.none()

    def get(self, request):
        hall_id = request.GET.get('hall', '')
        tmpl_id = request.GET.get('template', '')
        halls = self.get_allowed_halls(request)
        templates = ScheduleTemplate.objects.filter(is_active=True)

        if hall_id:
            halls = halls.filter(pk=hall_id)

        if tmpl_id:
            halls = halls.filter(schedule_template_id=tmpl_id)

        colors = ['sc-0', 'sc-1', 'sc-2', 'sc-3', 'sc-4', 'sc-5', 'sc-6', 'sc-7']
        timetable = []

        for hall in halls:
            if hall.schedule_template:
                entries = hall.schedule_template.entries.select_related('subject').order_by('day', 'start_time')
            else:
                entries = hall.schedules.select_related('subject').order_by('day', 'start_time')

            days_map = {}

            for e in entries:
                days_map.setdefault(e.day, []).append(e)

            days_list = []

            for day_key, day_label in ALL_DAYS:
                sessions = days_map.get(day_key, [])

                for i, s in enumerate(sessions):
                    s.color = colors[i % len(colors)]

                days_list.append({
                    'key': day_key,
                    'label': day_label,
                    'sessions': sessions
                })

            active_days = [d for d in days_list if d['sessions']]
            total = sum(len(d['sessions']) for d in days_list)

            timetable.append({
                'hall': hall,
                'days': days_list,
                'active_days': active_days,
                'total_sch': total,
                'tmpl': hall.schedule_template,
            })

        return render(request, 'halls/all_schedules.html', {
            'timetable': timetable,
            'all_halls': self.get_allowed_halls(request),
            'all_templates': templates,
            'selected_hall': hall_id,
            'selected_tmpl': tmpl_id,
        })


# ══════════════════════════════════════════
#  المواد الدراسية
# ══════════════════════════════════════════

class SubjectListView(GeneralSupervisorRequiredMixin, View):
    def get(self, request):
        subjects = Subject.objects.all()

        return render(request, 'halls/subjects/list.html', {
            'subjects': subjects,
            'total': subjects.count(),
            'active': subjects.filter(is_active=True).count(),
        })


class SubjectCreateView(GeneralSupervisorRequiredMixin, View):
    template_name = 'halls/subjects/form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'form': SubjectForm(),
            'action': 'إضافة'
        })

    def post(self, request):
        form = SubjectForm(request.POST)

        if form.is_valid():
            subject = form.save()

            messages.success(request, f'✅ تم إضافة مادة {subject.name} بنجاح')
            return redirect('halls:subjects')

        return render(request, self.template_name, {
            'form': form,
            'action': 'إضافة'
        })


class SubjectUpdateView(GeneralSupervisorRequiredMixin, View):
    template_name = 'halls/subjects/form.html'

    def get(self, request, pk):
        subject = get_object_or_404(Subject, pk=pk)

        return render(request, self.template_name, {
            'form': SubjectForm(instance=subject),
            'action': 'تعديل',
            'subject': subject
        })

    def post(self, request, pk):
        subject = get_object_or_404(Subject, pk=pk)
        form = SubjectForm(request.POST, instance=subject)

        if form.is_valid():
            form.save()

            messages.success(request, f'✅ تم تعديل مادة {subject.name} بنجاح')
            return redirect('halls:subjects')

        return render(request, self.template_name, {
            'form': form,
            'action': 'تعديل',
            'subject': subject
        })


class SubjectDeleteView(GeneralSupervisorRequiredMixin, View):
    def post(self, request, pk):
        subject = get_object_or_404(Subject, pk=pk)
        name = subject.name

        if subject.hallschedule_set.exists():
            messages.error(request, f'❌ لا يمكن حذف "{name}" — مستخدمة في جداول القاعات')
        else:
            subject.delete()
            messages.success(request, f'✅ تم حذف مادة {name}')

        return redirect('halls:subjects')


# ══════════════════════════════════════════
#  الجداول النموذجية ScheduleTemplate
# ══════════════════════════════════════════

class ScheduleTemplateListView(GeneralSupervisorRequiredMixin, View):
    def get(self, request):
        templates = ScheduleTemplate.objects.prefetch_related('entries', 'halls')

        return render(request, 'halls/tmpl/list.html', {
            'templates': templates,
            'total': templates.count(),
        })


class ScheduleTemplateCreateView(GeneralSupervisorRequiredMixin, View):
    template_name = 'halls/tmpl/form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'form': ScheduleTemplateForm(),
            'action': 'إنشاء'
        })

    def post(self, request):
        form = ScheduleTemplateForm(request.POST)

        if form.is_valid():
            tmpl = form.save()

            messages.success(request, f'✅ تم إنشاء جدول "{tmpl.name}"')
            return redirect('halls:template_detail', pk=tmpl.pk)

        return render(request, self.template_name, {
            'form': form,
            'action': 'إنشاء'
        })


class ScheduleTemplateDetailView(GeneralSupervisorRequiredMixin, View):
    template_name = 'halls/tmpl/detail.html'

    def get(self, request, pk):
        tmpl = get_object_or_404(ScheduleTemplate, pk=pk)
        entries = tmpl.entries.select_related('subject').order_by('day', 'start_time')

        days_map = {}

        for entry in entries:
            days_map.setdefault(entry.day, []).append(entry)

        days_data = [
            {
                'key': day_key,
                'label': day_label,
                'sessions': days_map[day_key]
            }
            for day_key, day_label in ALL_DAYS
            if day_key in days_map
        ]

        # القاعات المتاحة للمستخدم
        if request.user.is_general_manager:
            all_halls = Hall.objects.filter(
                is_active=True
            ).select_related('age_group')

        elif request.user.is_general_supervisor:
            all_halls = Hall.objects.filter(
                is_active=True,
                general_supervisor_assignments__supervisor=request.user
            ).select_related('age_group').distinct()

        else:
            all_halls = Hall.objects.none()

        # القاعات المرتبطة بهذا الجدول حالياً
        assigned_hall_ids = set(tmpl.halls.values_list('pk', flat=True))

        return render(request, self.template_name, {
            'tmpl': tmpl,
            'days_data': days_data,
            'halls': tmpl.halls.all(),
            'entry_form': ScheduleTemplateEntryForm(),
            'all_days': ALL_DAYS,
            'subjects': Subject.objects.filter(is_active=True),
            'all_halls': all_halls,
            'assigned_hall_ids': assigned_hall_ids,
        })


class ScheduleTemplateUpdateView(GeneralSupervisorRequiredMixin, View):
    template_name = 'halls/tmpl/form.html'

    def get(self, request, pk):
        tmpl = get_object_or_404(ScheduleTemplate, pk=pk)

        return render(request, self.template_name, {
            'form': ScheduleTemplateForm(instance=tmpl),
            'action': 'تعديل',
            'tmpl': tmpl
        })

    def post(self, request, pk):
        tmpl = get_object_or_404(ScheduleTemplate, pk=pk)
        form = ScheduleTemplateForm(request.POST, instance=tmpl)

        if form.is_valid():
            form.save()

            messages.success(request, f'✅ تم تعديل جدول "{tmpl.name}"')
            return redirect('halls:template_detail', pk=tmpl.pk)

        return render(request, self.template_name, {
            'form': form,
            'action': 'تعديل',
            'tmpl': tmpl
        })


class ScheduleTemplateDeleteView(GeneralSupervisorRequiredMixin, View):
    def post(self, request, pk):
        tmpl = get_object_or_404(ScheduleTemplate, pk=pk)

        if tmpl.halls.exists():
            messages.error(
                request,
                f'❌ لا يمكن حذف "{tmpl.name}" — مُسند لـ {tmpl.halls.count()} قاعة'
            )
        else:
            name = tmpl.name
            tmpl.delete()
            messages.success(request, f'✅ تم حذف جدول "{name}"')

        return redirect('halls:templates')


class TemplateEntryAddView(GeneralSupervisorRequiredMixin, View):
    def post(self, request, pk):
        tmpl = get_object_or_404(ScheduleTemplate, pk=pk)
        form = ScheduleTemplateEntryForm(request.POST)

        if form.is_valid():
            entry = form.save(commit=False)
            entry.template = tmpl
            entry.save()

            messages.success(request, f'✅ تمت إضافة الحصة ليوم {entry.get_day_display()}')

        else:
            for field, errors in form.errors.items():
                for err in errors:
                    messages.error(request, f'❌ {err}')

        return redirect('halls:template_detail', pk=pk)


class TemplateEntryDeleteView(GeneralSupervisorRequiredMixin, View):
    def post(self, request, entry_pk):
        entry = get_object_or_404(ScheduleTemplateEntry, pk=entry_pk)
        tmpl_pk = entry.template_id
        entry.delete()

        messages.success(request, '✅ تم حذف الحصة')
        return redirect('halls:template_detail', pk=tmpl_pk)


class TemplateAssignHallsView(GeneralSupervisorRequiredMixin, View):
    """
    إسناد / إلغاء إسناد جدول نموذجي من/إلى قاعات.
    تم تعديلها لتدعم القاعات المسندة للمشرف العام من جدول GeneralSupervisorHallAssignment.
    """
    def post(self, request, pk):
        tmpl = get_object_or_404(ScheduleTemplate, pk=pk)
        selected = set(request.POST.getlist('halls'))

        if request.user.is_general_manager:
            allowed = Hall.objects.filter(is_active=True)

        elif request.user.is_general_supervisor:
            allowed = Hall.objects.filter(
                is_active=True,
                general_supervisor_assignments__supervisor=request.user
            ).distinct()

        else:
            allowed = Hall.objects.none()

        for hall in allowed:
            if str(hall.pk) in selected:
                if hall.schedule_template != tmpl:
                    hall.schedule_template = tmpl
                    hall.save(update_fields=['schedule_template'])
            else:
                if hall.schedule_template == tmpl:
                    hall.schedule_template = None
                    hall.save(update_fields=['schedule_template'])

        messages.success(request, '✅ تم تحديث إسناد الجدول للقاعات')
        return redirect('halls:template_detail', pk=pk)


class HallExportExcelView(StaffRequiredMixin, View):
    def get(self, request, pk):
        hall = get_object_or_404(
            Hall.objects.select_related(
                'teacher',
                'supervisor',
                'age_group',
            ).prefetch_related(
                'general_supervisor_assignments__supervisor'
            ),
            pk=pk
        )

        if not user_can_access_hall(request.user, hall):
            messages.error(request, 'ليس لديك صلاحية تصدير بيانات هذه القاعة')
            return redirect('halls:list')

        students = hall.students.filter(
            status='active'
        ).select_related(
            'parent'
        ).order_by(
            'first_name',
            'last_name'
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "بيانات القاعة"

        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        header_fill = PatternFill("solid", fgColor="01696F")
        sub_fill = PatternFill("solid", fgColor="D9EDEB")

        title_font = Font(bold=True, size=16, color="FFFFFF")
        header_font = Font(bold=True, color="FFFFFF")
        bold_font = Font(bold=True)

        center = Alignment(horizontal='center', vertical='center')
        right = Alignment(horizontal='right', vertical='center')

        # عنوان التقرير
        ws.merge_cells('A1:F1')
        ws['A1'] = f'كشف طلاب القاعة: {hall.name}'
        ws['A1'].font = title_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center

        # معلومات القاعة
        info_rows = [
            ('اسم القاعة', hall.name),
            ('الموقع', hall.location or '—'),
            ('الفئة العمرية', hall.age_group.name if hall.age_group else '—'),
            ('المعلم', hall.teacher.get_full_name() if hall.teacher else '—'),
            ('مشرف القاعة', hall.supervisor.get_full_name() if hall.supervisor else '—'),
            ('المشرفون العامون', get_general_supervisor_names(hall)),
            ('العدد الحالي', students.count()),
        ]

        row = 3

        for label, value in info_rows:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = bold_font
            ws[f'A{row}'].fill = sub_fill
            ws[f'A{row}'].border = border
            ws[f'A{row}'].alignment = right

            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=2).alignment = right

            for c in range(2, 7):
                ws.cell(row=row, column=c).border = border

            row += 1

        row += 1

        # عنوان جدول الطلاب
        headers = ['م', 'اسم الطالب', 'ولي الأمر', 'الهاتف', 'تاريخ الميلاد', 'الحالة']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # بيانات الطلاب
        for index, student in enumerate(students, start=1):
            row += 1

            ws.cell(row=row, column=1, value=index).alignment = center
            ws.cell(
                row=row,
                column=2,
                value=student.get_full_name()
                if hasattr(student, 'get_full_name')
                else f"{student.first_name} {student.last_name}"
            )
            ws.cell(
                row=row,
                column=3,
                value=student.parent.get_full_name() if student.parent else '—'
            )
            ws.cell(
                row=row,
                column=4,
                value=getattr(student.parent, 'phone', '—') if student.parent else '—'
            )
            ws.cell(
                row=row,
                column=5,
                value=student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '—'
            )
            ws.cell(
                row=row,
                column=6,
                value=student.get_status_display()
                if hasattr(student, 'get_status_display')
                else student.status
            )

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = right if col != 1 else center

        # عرض الأعمدة
        widths = {
            1: 8,
            2: 28,
            3: 28,
            4: 18,
            5: 18,
            6: 14,
        }

        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = f"hall_{hall.pk}_students.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response